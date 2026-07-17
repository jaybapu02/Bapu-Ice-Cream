import csv
import io
from decimal import Decimal
from datetime import datetime

from django.contrib import admin
from django.contrib.auth.models import User
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin, GroupAdmin as BaseGroupAdmin
from django.contrib.auth.models import Group
from django.http import HttpResponse
from django.utils.html import format_html
from django.urls import reverse
from django.db.models import Q, Count

from .models import (
    Contact, CateringEnquiry, Order, OrderItem,
    Product, Category, Review, Newsletter, Wishlist
)
from .cadmin import custom_admin_site

# ──────────────────────────────────────────────
# Custom Admin Actions (Export)
# ──────────────────────────────────────────────

def export_as_csv(modeladmin, request, queryset):
    opts = modeladmin.model._meta
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={opts.verbose_name_plural}.csv'
    writer = csv.writer(response)
    field_names = [field.name for field in opts.fields]
    writer.writerow(field_names)
    for obj in queryset:
        writer.writerow([getattr(obj, field) for field in field_names])
    return response
export_as_csv.short_description = "📄 Export Selected as CSV"


def export_as_excel(modeladmin, request, queryset):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        modeladmin.message_user(request, "openpyxl is not installed. Install it with: pip install openpyxl")
        return

    opts = modeladmin.model._meta
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = opts.verbose_name_plural

    field_names = [field.name for field in opts.fields]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")

    for col_idx, field_name in enumerate(field_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=field_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, obj in enumerate(queryset, 2):
        for col_idx, field_name in enumerate(field_names, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(getattr(obj, field_name)))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={opts.verbose_name_plural}.xlsx'
    wb.save(response)
    return response
export_as_excel.short_description = "📊 Export Selected as Excel"


def export_as_pdf(modeladmin, request, queryset):
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
    except ImportError:
        modeladmin.message_user(request, "reportlab is not installed. Install it with: pip install reportlab")
        return

    opts = modeladmin.model._meta
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={opts.verbose_name_plural}.pdf'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    title = f"{opts.verbose_name_plural}".title()
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 0.25 * inch))

    field_names = [field.name for field in opts.fields]
    data = [field_names]
    for obj in queryset:
        data.append([str(getattr(obj, field)) for field in field_names])

    col_width = max(1.5 * inch, (10 * inch) / len(field_names))
    table = Table(data, colWidths=[col_width] * len(field_names))

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ])
    table.setStyle(style)
    elements.append(table)
    doc.build(elements)
    return response
export_as_pdf.short_description = "📑 Export Selected as PDF"


# ──────────────────────────────────────────────
# Inline Classes
# ──────────────────────────────────────────────

class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ["created_at"]
    fields = ["product", "ice_cream_type", "flavour", "size", "toppings", "quantity", "price", "created_at"]
    autocomplete_fields = ["product"]
    classes = ["collapse"]
    show_change_link = True


class ReviewInline(admin.TabularInline):
    model = Review
    extra = 0
    readonly_fields = ["created_at"]
    fields = ["name", "rating", "comment", "created_at"]
    classes = ["collapse"]
    can_delete = False
    verbose_name_plural = "Reviews"
    max_num = 10


class WishlistInline(admin.TabularInline):
    model = Wishlist
    extra = 0
    readonly_fields = ["created_at"]
    fields = ["product", "created_at"]
    autocomplete_fields = ["product"]
    classes = ["collapse"]
    can_delete = False
    verbose_name_plural = "Wishlist Items"


# ──────────────────────────────────────────────
# Mixin — shared image thumbnail helper
# ──────────────────────────────────────────────

class ImageThumbnailMixin:
    def thumbnail(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" style="width:50px;height:50px;object-fit:cover;'
                'border-radius:6px;border:1px solid #dee2e6;" />',
                obj.image.url
            )
        return format_html(
            '<span style="color:#6c757d;font-size:1.5rem;">📷</span>'
        )
    thumbnail.short_description = "Image"


# ──────────────────────────────────────────────
# Category Admin
# ──────────────────────────────────────────────

@admin.register(Category, site=custom_admin_site)
class CategoryAdmin(ImageThumbnailMixin, admin.ModelAdmin):
    list_display = ["thumbnail", "name", "slug", "products_count", "created_at"]
    list_display_links = ["thumbnail", "name"]
    search_fields = ["name", "description"]
    prepopulated_fields = {"slug": ("name",)}
    readonly_fields = ["created_at", "products_count"]
    list_per_page = 20
    ordering = ["name"]
    fieldsets = [
        (None, {"fields": ["name", "slug", "description"]}),
        ("Image", {"fields": ["image", "thumbnail"], "classes": ["collapse"]}),
        ("Timestamps", {"fields": ["created_at"], "classes": ["collapse"]}),
    ]

    def products_count(self, obj):
        count = getattr(obj, 'prod_count', obj.products.count())
        url = reverse("admin:home_product_changelist") + f"?category__id__exact={obj.id}"
        return format_html('<a href="{}" class="badge bg-primary">{}</a>', url, count)
    products_count.short_description = "Products"

    def get_queryset(self, request):
        return super().get_queryset(request).annotate(prod_count=Count("products"))


# ──────────────────────────────────────────────
# Product Admin
# ──────────────────────────────────────────────

@admin.register(Product, site=custom_admin_site)
class ProductAdmin(ImageThumbnailMixin, admin.ModelAdmin):
    list_display = [
        "thumbnail", "name", "category", "price_display",
        "stock_status", "is_available", "is_featured", "created_at"
    ]
    list_display_links = ["thumbnail", "name"]
    list_filter = ["category", "is_available", "is_featured", "created_at"]
    search_fields = ["name", "description", "category__name"]
    prepopulated_fields = {"slug": ("name",)}
    list_editable = ["is_available", "is_featured"]
    list_per_page = 25
    ordering = ["-created_at"]
    date_hierarchy = "created_at"
    save_on_top = True
    actions = [export_as_csv, export_as_excel, export_as_pdf]

    fieldsets = [
        (None, {
            "fields": ["category", "name", "slug", "description"]
        }),
        ("Pricing & Stock", {
            "fields": ["price", "stock", "is_available", "is_featured"],
            "classes": ["collapse"]
        }),
        ("Image", {
            "fields": ["image", "thumbnail"],
            "classes": ["collapse"]
        }),
        ("Timestamps", {
            "fields": ["created_at", "updated_at"],
            "classes": ["collapse"]
        }),
    ]

    readonly_fields = ["created_at", "updated_at", "thumbnail"]
    autocomplete_fields = ["category"]

    def price_display(self, obj):
        return format_html(
            '<span style="font-weight:600;color:#0d6efd;">₹{}</span>', obj.price
        )
    price_display.short_description = "Price"
    price_display.admin_order_field = "price"

    def stock_status(self, obj):
        if obj.stock == 0:
            return format_html('<span class="badge bg-danger">Out of Stock</span>')
        elif obj.stock < 5:
            return format_html(
                '<span class="badge bg-warning text-dark">Low ({})</span>', obj.stock
            )
        return format_html(
            '<span class="badge bg-success">In Stock ({})</span>', obj.stock
        )
    stock_status.short_description = "Stock"
    stock_status.admin_order_field = "stock"

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("category")

    class Media:
        css = {"all": ("admin/css/base.css",)}


# ──────────────────────────────────────────────
# Order Admin
# ──────────────────────────────────────────────

@admin.register(Order, site=custom_admin_site)
class OrderAdmin(admin.ModelAdmin):
    list_display = [
        "order_id", "name", "phone", "total_price_display",
        "payment_mode", "order_status_badge", "items_count", "created_at"
    ]
    list_filter = ["status", "payment_mode", "delivery_type", "created_at"]
    search_fields = ["order_id", "name", "phone", "address"]
    list_per_page = 25
    ordering = ["-created_at"]
    date_hierarchy = "created_at"
    save_on_top = True
    inlines = [OrderItemInline]
    actions = [export_as_csv, export_as_excel, export_as_pdf]

    fieldsets = [
        (None, {
            "fields": ["order_id", "status", "total_price"]
        }),
        ("Customer Details", {
            "fields": ["user", "name", "phone", "address"],
            "classes": ["collapse"]
        }),
        ("Delivery & Payment", {
            "fields": ["delivery_type", "payment_mode"],
            "classes": ["collapse"]
        }),
        ("Timestamps", {
            "fields": ["created_at", "updated_at"],
            "classes": ["collapse"]
        }),
    ]

    readonly_fields = ["order_id", "created_at", "updated_at"]
    autocomplete_fields = ["user"]

    def total_price_display(self, obj):
        return format_html(
            '<span style="font-weight:600;color:#0d6efd;">₹{}</span>', obj.total_price
        )
    total_price_display.short_description = "Total"
    total_price_display.admin_order_field = "total_price"

    def order_status_badge(self, obj):
        colors = {"PENDING": "warning", "PAID": "success", "CANCELLED": "danger"}
        badge_class = colors.get(obj.status, "secondary")
        return format_html(
            '<span class="badge bg-{}">{}</span>', badge_class, obj.get_status_display()
        )
    order_status_badge.short_description = "Status"
    order_status_badge.admin_order_field = "status"

    def items_count(self, obj):
        count = obj.items.count()
        return format_html(
            '<span class="badge bg-info">{}</span>', count
        )
    items_count.short_description = "Items"

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("user").prefetch_related("items")


# ──────────────────────────────────────────────
# OrderItem Admin
# ──────────────────────────────────────────────

@admin.register(OrderItem, site=custom_admin_site)
class OrderItemAdmin(admin.ModelAdmin):
    list_display = ["order_link", "product", "ice_cream_type", "flavour", "size", "quantity", "price_display", "created_at"]
    list_filter = ["flavour", "ice_cream_type", "created_at"]
    search_fields = ["flavour", "order__order_id", "product__name"]
    list_per_page = 25
    ordering = ["-created_at"]
    readonly_fields = ["created_at"]
    autocomplete_fields = ["order", "product"]

    def order_link(self, obj):
        url = reverse("admin:home_order_change", args=[obj.order.id])
        return format_html('<a href="{}">{}</a>', url, obj.order.order_id)
    order_link.short_description = "Order"
    order_link.admin_order_field = "order__order_id"

    def price_display(self, obj):
        return format_html('₹{}', obj.price)
    price_display.short_description = "Price"
    price_display.admin_order_field = "price"

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("order", "product")


# ──────────────────────────────────────────────
# Review Admin
# ──────────────────────────────────────────────

@admin.register(Review, site=custom_admin_site)
class ReviewAdmin(admin.ModelAdmin):
    list_display = ["product_link", "name", "rating_stars", "email", "comment_short", "created_at"]
    list_filter = ["rating", "created_at"]
    search_fields = ["name", "email", "comment", "product__name"]
    list_per_page = 25
    ordering = ["-created_at"]
    date_hierarchy = "created_at"
    readonly_fields = ["created_at"]
    autocomplete_fields = ["product", "user"]

    fieldsets = [
        (None, {
            "fields": ["product", "user", "name", "email"]
        }),
        ("Review", {
            "fields": ["rating", "comment"]
        }),
        ("Timestamps", {
            "fields": ["created_at"],
            "classes": ["collapse"]
        }),
    ]

    def product_link(self, obj):
        url = reverse("admin:home_product_change", args=[obj.product.id])
        return format_html('<a href="{}">{}</a>', url, obj.product.name[:50])
    product_link.short_description = "Product"
    product_link.admin_order_field = "product__name"

    def rating_stars(self, obj):
        full = "★" * obj.rating
        empty = "☆" * (5 - obj.rating)
        return format_html(
            '<span style="color:#ffc107;font-size:1.1rem;">{}{}</span>',
            full, empty
        )
    rating_stars.short_description = "Rating"
    rating_stars.admin_order_field = "rating"

    def comment_short(self, obj):
        return obj.comment[:75] + "..." if len(obj.comment) > 75 else obj.comment
    comment_short.short_description = "Comment"

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("product", "user")


# ──────────────────────────────────────────────
# Contact Admin
# ──────────────────────────────────────────────

@admin.register(Contact, site=custom_admin_site)
class ContactAdmin(admin.ModelAdmin):
    list_display = ["name", "email", "phone", "message_short", "created_at"]
    list_filter = ["created_at"]
    search_fields = ["name", "email", "phone", "message"]
    list_per_page = 25
    ordering = ["-created_at"]
    date_hierarchy = "created_at"
    readonly_fields = ["created_at", "updated_at"]
    actions = [export_as_csv, export_as_excel, export_as_pdf]

    fieldsets = [
        (None, {
            "fields": ["name", "email", "phone"]
        }),
        ("Message", {
            "fields": ["message"]
        }),
        ("Timestamps", {
            "fields": ["created_at", "updated_at"],
            "classes": ["collapse"]
        }),
    ]

    def message_short(self, obj):
        return obj.message[:75] + "..." if len(obj.message) > 75 else obj.message
    message_short.short_description = "Message"


# ──────────────────────────────────────────────
# CateringEnquiry Admin
# ──────────────────────────────────────────────

@admin.register(CateringEnquiry, site=custom_admin_site)
class CateringEnquiryAdmin(admin.ModelAdmin):
    list_display = [
        "name", "phone", "email", "event_type", "event_date",
        "venue", "guests", "catering_package", "budget_display",
        "reference_thumbnail", "created_at"
    ]
    list_filter = ["event_type", "catering_package", "event_date", "created_at"]
    search_fields = ["name", "phone", "email", "event_type", "venue", "message"]
    list_per_page = 25
    ordering = ["-created_at"]
    date_hierarchy = "created_at"
    readonly_fields = ["created_at", "updated_at", "reference_thumbnail"]
    actions = [export_as_csv, export_as_excel, export_as_pdf]

    fieldsets = [
        (None, {
            "fields": ["name", "phone", "email"]
        }),
        ("Event Details", {
            "fields": ["event_type", "event_date", "venue", "guests"]
        }),
        ("Package & Budget", {
            "fields": ["catering_package", "budget"],
            "classes": ["collapse"]
        }),
        ("Requirements", {
            "fields": ["special_requirements", "message", "reference_image", "reference_thumbnail"],
            "classes": ["collapse"]
        }),
        ("Timestamps", {
            "fields": ["created_at", "updated_at"],
            "classes": ["collapse"]
        }),
    ]

    def budget_display(self, obj):
        if obj.budget:
            return format_html(
                '<span style="font-weight:600;color:#0d6efd;">₹{}</span>', obj.budget
            )
        return format_html('<span class="text-muted">—</span>')
    budget_display.short_description = "Budget"
    budget_display.admin_order_field = "budget"

    def reference_thumbnail(self, obj):
        if obj.reference_image:
            return format_html(
                '<a href="{}" target="_blank">'
                '<img src="{}" style="width:60px;height:60px;object-fit:cover;'
                'border-radius:6px;border:1px solid #dee2e6;" /></a>',
                obj.reference_image.url, obj.reference_image.url
            )
        return format_html('<span class="text-muted">—</span>')
    reference_thumbnail.short_description = "Reference Image"


# ──────────────────────────────────────────────
# Newsletter Admin
# ──────────────────────────────────────────────

@admin.register(Newsletter, site=custom_admin_site)
class NewsletterAdmin(admin.ModelAdmin):
    list_display = ["email", "is_active_badge", "created_at"]
    list_filter = ["is_active", "created_at"]
    search_fields = ["email"]
    list_per_page = 25
    ordering = ["-created_at"]
    readonly_fields = ["created_at"]
    actions = [export_as_csv, export_as_excel, export_as_pdf]

    fieldsets = [
        (None, {
            "fields": ["email", "is_active"]
        }),
        ("Timestamps", {
            "fields": ["created_at"],
            "classes": ["collapse"]
        }),
    ]

    def is_active_badge(self, obj):
        if obj.is_active:
            return format_html('<span class="badge bg-success">Active</span>')
        return format_html('<span class="badge bg-secondary">Inactive</span>')
    is_active_badge.short_description = "Status"
    is_active_badge.admin_order_field = "is_active"


# ──────────────────────────────────────────────
# Wishlist Admin
# ──────────────────────────────────────────────

@admin.register(Wishlist, site=custom_admin_site)
class WishlistAdmin(admin.ModelAdmin):
    list_display = ["user", "product_link", "created_at"]
    search_fields = ["user__username", "user__email", "product__name"]
    list_filter = ["created_at"]
    list_per_page = 25
    ordering = ["-created_at"]
    readonly_fields = ["created_at"]
    autocomplete_fields = ["user", "product"]

    def product_link(self, obj):
        url = reverse("admin:home_product_change", args=[obj.product.id])
        return format_html('<a href="{}">{}</a>', url, obj.product.name[:50])
    product_link.short_description = "Product"
    product_link.admin_order_field = "product__name"

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("user", "product")


# ──────────────────────────────────────────────
# User & Group Admin (register with custom site)
# ──────────────────────────────────────────────

class CustomUserAdmin(BaseUserAdmin):
    list_display = ["username", "email", "first_name", "last_name", "is_staff", "is_active", "date_joined"]
    list_filter = ["is_staff", "is_active", "is_superuser", "groups"]
    search_fields = ["username", "email", "first_name", "last_name"]
    list_per_page = 25
    ordering = ["-date_joined"]
    fieldsets = [
        (None, {"fields": ["username", "password"]}),
        ("Personal Info", {"fields": ["first_name", "last_name", "email"], "classes": ["collapse"]}),
        ("Permissions", {
            "fields": ["is_active", "is_staff", "is_superuser", "groups", "user_permissions"],
            "classes": ["collapse"]
        }),
        ("Important Dates", {"fields": ["last_login", "date_joined"], "classes": ["collapse"]}),
    ]
    filter_horizontal = ["groups", "user_permissions"]


class CustomGroupAdmin(BaseGroupAdmin):
    list_display = ["name", "user_count"]
    search_fields = ["name"]
    ordering = ["name"]
    filter_horizontal = ["permissions"]

    def user_count(self, obj):
        count = obj.user_set.count()
        return format_html('<span class="badge bg-primary">{}</span>', count)
    user_count.short_description = "Users"


custom_admin_site.register(User, CustomUserAdmin)
custom_admin_site.register(Group, CustomGroupAdmin)
